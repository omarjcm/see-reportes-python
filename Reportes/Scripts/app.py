import openpyxl
# Esto es para acceder a la fecha y hora del computador
from datetime import datetime
from report_excel import *
from data import *

from openpyxl.styles import numbers
import pandas as pd

now = datetime.now()
'''
 %b  -> Mes
 %d  -> Día
 %Y  -> Año
 %H  -> Hora
 %M  -> Minutos
 %S  -> Segundos
'''
date_now = now.strftime( "%b-%d-%Y_%H-%M-%S" )

nombre_archivo = '../Data/reporte_see' + date_now + '.xlsx'
wb = openpyxl.Workbook()
sheet = wb['Sheet']
sheet.title = 'Reporte 1'

agregar_imagen( sheet )

ws = wb.active
ws.oddHeader.left.color = 'CC3366'

agregar_titulo( ws )

cabecera = [
        'Country', 
        'Year', 
        '1', '2', '3', '4', '5', '6', '7', '8', '9', '10', '11', '12']

agregar_filtro(ws, sheet, cabecera)

# df_paises -> ventas por pais y año.
# df_transacciones -> ventas por pais, por año y por mes.
df_paises, df_transacciones = get_transacciones_x_pais()

# La fila donde va a iniciar la inserción de la data en excel.
inicial = 8
final = len( df_paises ) + inicial
sheet.insert_rows( inicial, final )

for index_row, data_row in df_paises.iterrows():
    index_col = 1
    data = list( data_row )
    
    sheet.cell( row=inicial, column=index_col, value=data_row['Country'] )
    index_col += 1
    sheet.cell( row=inicial, column=index_col, value=data_row['anio'] )
    index_col += 1
    
    cabecera_index = 2
    while( cabecera_index < len( cabecera ) ):
        loc_country = df_transacciones['Country'] == data_row['Country']
        loc_anio = df_transacciones['anio'] == data_row['anio']
        loc_mes = df_transacciones['mes'] == int(cabecera[ cabecera_index ])
        
        ventas_x_mes = df_transacciones.loc[ loc_country & loc_anio & loc_mes ]
        
        cell = None
        if len(ventas_x_mes) > 0:
            cell = sheet.cell( row=inicial, column=index_col, value=ventas_x_mes['venta_usd'].values[0] )
        else:
            cell = sheet.cell( row=inicial, column=index_col, value='' )
        
        cell.number_format = numbers.FORMAT_CURRENCY_USD
        cabecera_index += 1
        index_col += 1
        
    inicial += 1

wb.save(nombre_archivo)








