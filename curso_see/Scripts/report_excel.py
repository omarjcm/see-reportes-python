import openpyxl
import pandas
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment
from os import path

thin = Side(border_style="thin", color="0089bb")
titulo_border = Border(top=thin, left=thin, right=thin, bottom=thin)
titulo_rect = PatternFill("solid", fgColor="0089bb")
titulo_font_big = Font(name='Calibri', size=16, b=True, color="ffffff")
titulo_font_little = Font(name='Calibri', size=12, b=True, color="ffffff")
titulo_alignment = Alignment(horizontal="center", vertical="center")

def load_workbook( wb_path ):
    if path.exists( wb_path ):
        return openpyxl.load_workbook( wb_path )
    return openpyxl.Workbook()

def agregar_imagen( sheet ):
    img = openpyxl.drawing.image.Image('../Data/image/see_webinar.png')
    sheet.add_image(img, 'A1')

def agregar_titulo( ws ):
    ws.merge_cells('A5:N5')
    top_left_cell = ws['A5']
    top_left_cell.value = "Reporte de los Países"
    
    top_left_cell.border = titulo_border
    top_left_cell.fill = titulo_rect
    top_left_cell.font = titulo_font_big
    top_left_cell.alignment = titulo_alignment

def agregar_filtro(ws, sheet, cabecera):
    cells = sheet['A7:N7']
    indice = 0
    for row in cells:
        for cell in row:
            cell.value = cabecera[indice]
            cell.fill = titulo_rect
            cell.font = titulo_font_little
            cell.alignment = titulo_alignment
            indice += 1

    ws.auto_filter.ref = 'A7:N7'
