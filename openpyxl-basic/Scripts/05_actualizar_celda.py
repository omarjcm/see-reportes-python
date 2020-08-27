import openpyxl
from funciones import *

ruta = '../Data/test.xlsx'
wb = cargar_archivo( ruta )
hoja = obtener_hoja(wb, 'Habitos')

hoja.cell( row=1, column=1, value='Num. de Serie' )

hoja.cell( row=1, column=1 ).value = 'Serie Num.'

celda = hoja.cell( row=1, column=1 )
celda.value = 'NÃºmero de Serie'
celda.font = openpyxl.styles.Font(bold=True, color='0362fc')

wb.save( ruta )