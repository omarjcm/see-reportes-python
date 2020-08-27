import openpyxl
from os import path

# Esto para definir una función en Python
def cargar_archivo( ruta_archivo ):
    if path.exists( ruta_archivo ):
        return openpyxl.load_workbook( ruta_archivo )
    else:
        return openpyxl.Workbook()
    
def obtener_hoja(wb, titulo_hoja):
    if wb.sheetnames[0] != titulo_hoja:
        hoja = wb['Sheet']
        hoja.title = titulo_hoja
    else:
        hoja = wb[titulo_hoja]

    return hoja
